"""
Script to load initial researcher data from Excel files into the database.
Run this once after deploying to populate the database with existing researchers.
"""

from app import app, db, Researcher
import openpyxl

def load_external_researchers(filename):
    """Load external researchers from Excel file"""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        
        name = data.get('Your Name', '')
        email = data.get('Email Address', '')
        
        if not name or not email:
            continue
        
        researcher = Researcher(
            name=name,
            email=email,
            organization=data.get('Your Orgnization') or data.get('Your Organization', ''),
            researcher_type='external',
            organization_focus=data.get('What is your organization\'s primary area of focus or industry sector?Please list key words or phrases (e.g., renewable energy, healthcare, logistics, education technology)', ''),
            challenge_description=data.get('Please describe a challenge or business goal your organization is currently facing that could benefit from academic collaboration.\n(e.g., improving supply chain efficiency, developing sustainable mate', ''),
            expertise_sought=data.get('What type of expertise or research support are you seeking to address this challenge?(e.g., machine learning, food security, sustainable packaging, behavioral economics)', ''),
            lab_tours_interested=data.get('Which lab tour(s) would you be interested in joining during our event? (Tour selection will be finalized at the event. As tour lengths will vary, it is anticipated that participants will have time to ', '')
        )
        
        try:
            db.session.add(researcher)
            db.session.commit()
            count += 1
        except Exception as e:
            db.session.rollback()
            print(f"Error adding {name}: {e}")
    
    return count

def load_internal_researchers_file1(filename):
    """Load internal researchers from first Excel file"""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        
        name = data.get('Your Name', '')
        email = data.get('Email Address', '')
        
        if not name or not email:
            continue
        
        researcher = Researcher(
            name=name,
            email=email,
            organization='Humber Polytechnic',
            researcher_type='internal',
            faculty_department=data.get('Your Faculty/Department', ''),
            primary_areas=data.get('What are your primary areas of research or expertise?Please list key words or phrases (e.g., machine learning, food security, sustainable packaging, behavioral economics).', ''),
            experience_summary=data.get('Please provide a brief summary of your experience or capabilities relevant to collaborative research?(e.g., summary of technical skills, related past work, specialized expertise)', ''),
            sectors_interested=data.get('What sectors or societal challenges are you most interested in addressing through research?(e.g., healthcare innovation, climate resilience, advanced manufacturing, education equity)', '')
        )
        
        try:
            db.session.add(researcher)
            db.session.commit()
            count += 1
        except Exception as e:
            db.session.rollback()
            print(f"Error adding {name}: {e}")
    
    return count

def load_internal_researchers_file2(filename):
    """Load additional internal researchers from second Excel file"""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        
        name = data.get('Researcher Full Name', '')
        email = data.get('Humber Email', '')
        
        if not name or not email:
            continue
        
        researcher = Researcher(
            name=name,
            email=email,
            organization='Humber Polytechnic',
            researcher_type='internal',
            faculty_department=data.get('Faculty / Department', ''),
            primary_areas=data.get('Research Interest Keywords', ''),
            experience_summary=data.get('Job Title', ''),
            sectors_interested=''
        )
        
        try:
            db.session.add(researcher)
            db.session.commit()
            count += 1
        except Exception as e:
            db.session.rollback()
            # Skip duplicates silently
            if 'UNIQUE constraint failed' not in str(e):
                print(f"Error adding {name}: {e}")
    
    return count

if __name__ == '__main__':
    with app.app_context():
        # Create tables
        db.create_all()
        print("Database tables created")
        
        # Load data (update these paths to your actual file locations)
        print("\nLoading external researchers...")
        ext_count = load_external_researchers('ResearchUnplugged_PartneringSessions!_ExternalResearcher.xlsx')
        print(f"Loaded {ext_count} external researchers")
        
        print("\nLoading internal researchers (File 1)...")
        int_count1 = load_internal_researchers_file1('ResearchUnplugged_PartneringSessions!1_InternalHumber.xlsx')
        print(f"Loaded {int_count1} internal researchers from File 1")
        
        print("\nLoading additional internal researchers (File 2)...")
        int_count2 = load_internal_researchers_file2('AdditonalInternalHumberResearcher.xlsx')
        print(f"Loaded {int_count2} additional internal researchers from File 2")
        
        # Print summary
        total_internal = Researcher.query.filter_by(researcher_type='internal').count()
        total_external = Researcher.query.filter_by(researcher_type='external').count()
        
        print("\n" + "="*50)
        print("DATABASE SUMMARY")
        print("="*50)
        print(f"Total Internal (Humber): {total_internal}")
        print(f"Total External: {total_external}")
        print(f"Total Researchers: {total_internal + total_external}")
        print("="*50)
